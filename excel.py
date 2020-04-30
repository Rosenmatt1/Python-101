import openpyxl as xl
from openpyxl import BarChart, Reference

def process_workbook(filename):
  wb = xl.load_workbook(filename)
  sheet = wb['Sheet1']
  print(sheet.max.row)

  # cell = sheet['a1']
  # cell = sheet.cell(1, 1)
  # print(cell.value)

  # starts at 2 becasue dont want header row
  for row in range(2, sheet.max.row + 1):
    print(row)
    cell = sheet.cell(row, 3)

    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell = corrected_price

  values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
  chart = BarChart()
  chart.add_data(values)
  sheet.add.chart(chart, 'e2')

  wb.save(filename)


